"""Shared code for working with workbook data.

Module was developed for sample sheet work but extracted for reuse with fetch
workbooks.
"""

import base64
from collections.abc import Generator
from csv import (
    Dialect,
    reader,
    Sniffer,
)
from dataclasses import dataclass
from io import (
    BytesIO,
    StringIO,
)
from textwrap import wrap
from typing import (
    Any,
    Literal,
    Optional,
    Protocol,
    Union,
)

from openpyxl import (
    load_workbook,
    Workbook,
)
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from galaxy.exceptions import RequestParameterInvalidException
from galaxy.model.dataset_collections.rule_target_models import (
    ColumnTarget,
)

Base64StringT = str


def freeze_header_row(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "A2"


def make_headers_bold(worksheet: Worksheet, headers: list) -> None:
    for index in range(len(headers)):
        make_bold(worksheet, 1, index)


def make_bold(worksheet: Worksheet, row: int, column: int):
    """Make the target cell bold in the specified worksheet."""
    bold_font = Font(bold=True)
    worksheet[f"{index_to_excel_column(column)}{row}"].font = bold_font


def index_to_excel_column(index: int) -> str:
    """Converts a numeric index (0-based) into an Excel column label."""
    if index < 0:
        raise ValueError("Index must be 0 or greater")

    column_label = ""
    while index >= 0:
        column_label = chr(index % 26 + 65) + column_label
        index = index // 26 - 1  # Move to the next "digit" in base-26, adjusting for zero-based indexing

    return column_label


def uri_data_validation(column: str) -> DataValidation:
    # Add data validation for "URI" column
    # We cannot assume http/https since drs, gxfiles, etc... are all fine
    uri_validation = DataValidation(type="custom", formula1=f'=ISNUMBER(FIND("://", {column}2))', allow_blank=True)
    uri_validation.error = "Invalid URI"
    uri_validation.errorTitle = "Error"
    uri_validation.showErrorMessage = True
    return uri_validation


def add_column_validation(column: str, data_validation: DataValidation, worksheet: Worksheet):
    data_validation.add(f"{column}2:{column}1048576")
    worksheet.add_data_validation(data_validation)


def set_column_width(worksheet: Worksheet, column_index: int, width: int):
    worksheet.column_dimensions[index_to_excel_column(column_index)].width = width


KnownContentTypes = Literal[
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/csv",
    "text/tab-separated-values",
    "text/plain",
]


class ReadOnlyWorkbook(Protocol):
    typed: bool
    content_type: KnownContentTypes

    def column_titles(self) -> list[str]:
        """Return the column titles from the first (or only) worksheet."""

    def iter_rows(self, num_columns: int) -> Generator[tuple[Any, ...], None, None]:
        """Iterate over the rows in the first (or only) worksheet, returning tuples of values."""


class ExcelReadOnlyWorkbook(ReadOnlyWorkbook):
    """A protocol for a read-only workbook, used to ensure compatibility with load_workbook."""

    typed: bool = True
    content_type: Literal["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    def __init__(self, workbook: Workbook):
        """Initialize the read-only workbook."""
        self._workbook = workbook

    def column_titles(self) -> list[str]:
        """Return the column titles from the first worksheet."""
        return read_column_header_titles(self._workbook.active)

    def iter_rows(self, num_columns: int) -> Generator[tuple[Any, ...], None, None]:
        """Iterate over the rows in the first worksheet, returning tuples of values."""
        return self._workbook.active.iter_rows(max_col=num_columns, values_only=True)


class CsvDialect(BaseModel):
    delimiter: str
    quote_character: Optional[str]
    double_quote: bool
    skip_initial_space: bool
    line_terminator: str
    escape_character: Optional[str]

    @staticmethod
    def from_csv_dialect(dialect: type[Dialect]) -> "CsvDialect":
        """Create a CsvDialect from a csv.Sniffer dialect."""
        return CsvDialect(
            delimiter=dialect.delimiter,
            quote_character=dialect.quotechar,
            double_quote=dialect.doublequote,
            skip_initial_space=dialect.skipinitialspace,
            line_terminator=dialect.lineterminator,
            escape_character=dialect.escapechar,
        )


class ContentTypeMessage(BaseModel):
    message: str
    content_type: str


class CsvDialectInferenceMessage(BaseModel):
    message: str
    dialect: CsvDialect


class CsvReaderReadOnlyWorkbook(ReadOnlyWorkbook):
    typed: bool = False
    content_type: KnownContentTypes

    def __init__(self, file_like: StringIO):
        """Initialize the read-only workbook."""
        self._file_like = file_like

        dialect = Sniffer().sniff(self._file_like.read(1024))
        self._file_like.seek(0)
        self._dialect = dialect
        if self._dialect.delimiter == "\t":
            self.content_type = "text/tab-separated-values"
        elif self._dialect.delimiter == ",":
            self.content_type = "text/csv"
        else:
            self.content_type = "text/plain"

    @property
    def dialect(self) -> CsvDialect:
        """Return the CSV dialect used by this workbook."""
        return CsvDialect.from_csv_dialect(self._dialect)

    def column_titles(self) -> list[str]:
        titles = next(reader(self._file_like, self._dialect))
        self._file_like.seek(0)  # Reset the file pointer after reading
        return titles

    def iter_rows(self, num_columns: int) -> Generator[tuple[Any, ...], None, None]:
        """Iterate over the rows in the CSV file, returning tuples of values."""
        csv_reader = reader(self._file_like, self._dialect)
        for row in csv_reader:
            if len(row) > num_columns:
                row = row[:num_columns]
            yield tuple(row)


def parse_format_messages(
    workbook: ReadOnlyWorkbook,
) -> list[Union[ContentTypeMessage, CsvDialectInferenceMessage]]:
    """Parse and return client-facing messages about parsing the target workbook."""
    messages: list[Union[ContentTypeMessage, CsvDialectInferenceMessage]] = []

    if isinstance(workbook, ExcelReadOnlyWorkbook):
        excel_content_type = workbook.content_type
        messages.append(
            ContentTypeMessage(
                message=f"This appears to be a xlsx workbook with content type {excel_content_type}.",
                content_type=workbook.content_type,
            )
        )
    elif isinstance(workbook, CsvReaderReadOnlyWorkbook):
        csv_content_type = workbook.content_type
        messages.append(
            ContentTypeMessage(
                message=f"This appears to be text workbook with content type {csv_content_type}.",
                content_type=workbook.content_type,
            )
        )
        # TODO: place CSV parsing parameters into the message in the future.
        messages.append(
            CsvDialectInferenceMessage(
                message="CSV parsing parameters we're inferred for this workbook.", dialect=workbook.dialect
            )
        )

    return messages


def load_workbook_from_base64(content: str) -> ReadOnlyWorkbook:
    decoded_content = base64.b64decode(content)
    file_like = BytesIO(decoded_content)
    is_excel = file_like.read(4) == b"\x50\x4b\x03\x04"
    workbook: ReadOnlyWorkbook
    file_like.seek(0)
    if is_excel:
        try:
            workbook = ExcelReadOnlyWorkbook(load_workbook(file_like, data_only=True))
        except Exception as e:
            extra_message = str(e)
            raise RequestParameterInvalidException(
                f"The provided content is not a valid Excel file (or at least not one Galaxy knows how to parse). Please check the content and try again. The underlying error was [{extra_message}]"
            )
    else:
        try:
            tabular = decoded_content.decode("utf-8")
            file_like_as_utf8 = StringIO(tabular)
            workbook = CsvReaderReadOnlyWorkbook(file_like_as_utf8)
        except Exception as e:
            extra_message = str(e)
            raise RequestParameterInvalidException(
                f"The provided content is not a parsable as a valid CSV or TSV (or at least not one Galaxy knows how to parse). Please check the content and try again. The underlying error was [{extra_message}]"
            )
    return workbook


def workbook_to_bytes(workbook: "Workbook") -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@dataclass
class HasHelp:
    title: str
    help: str


@dataclass
class HelpConfiguration:
    instructions: list[str]
    columns: list[HasHelp]
    text_width: int
    column_width: int
    help_row_start: int = 3


@dataclass
class ExtraColumnsHelpConfiguration:
    instructions: list[str]
    text_width: int
    column_targets: list[ColumnTarget]


def wrap_instructions(instruction: str, help_width: int) -> list[str]:
    return wrap(instruction, width=help_width)


def add_extra_column_help_as_new_sheet(workbook: Workbook, extra_columns_help: ExtraColumnsHelpConfiguration) -> None:
    worksheet = workbook.create_sheet(title="Supplying Extra Metadata")
    worksheet.cell(row=1, column=1, value="Metadata")
    worksheet.cell(row=1, column=2, value="Description")
    worksheet.cell(row=1, column=3, value="Example Column Names")
    make_bold(worksheet, 1, 0)
    make_bold(worksheet, 1, 1)
    make_bold(worksheet, 1, 2)

    set_column_width(worksheet, 0, 40)
    set_column_width(worksheet, 1, 140)
    set_column_width(worksheet, 2, 40)

    current_row = 2
    for column_target in extra_columns_help.column_targets:
        worksheet.cell(row=current_row, column=1, value=column_target.label)
        worksheet.cell(row=current_row, column=2, value=column_target.help)
        worksheet.cell(row=current_row, column=3, value=column_target.example_column_names_as_str)
        current_row += 1

    help_label_index = 6
    current_row = 2
    worksheet.cell(row=current_row, column=help_label_index, value="Instructions")
    make_bold(worksheet, current_row, help_label_index - 1)
    current_row += 1
    write_instructions_to_sheet(
        worksheet, extra_columns_help.instructions, current_row, help_label_index, extra_columns_help.text_width
    )


def add_instructions_to_sheet(worksheet: Worksheet, help_configuration: HelpConfiguration):
    columns = help_configuration.columns
    num_columns = len(columns)
    help_label_index = num_columns + 2
    # why is the width not +1
    set_column_width(worksheet, help_label_index, help_configuration.column_width)

    help_start_row = help_configuration.help_row_start
    current_row = help_start_row

    worksheet.cell(row=current_row, column=help_label_index, value="Instructions")
    make_bold(worksheet, current_row, help_label_index - 1)

    current_row += 1

    current_row = write_instructions_to_sheet(
        worksheet, help_configuration.instructions, current_row, help_label_index, help_configuration.text_width
    )

    current_row += 2
    worksheet.cell(row=current_row, column=help_label_index, value="Columns")
    make_bold(worksheet, current_row, help_label_index - 1)
    current_row += 1

    for column in columns:
        worksheet.cell(row=current_row, column=help_label_index, value=column.title)
        worksheet.cell(row=current_row, column=help_label_index + 1, value=column.help)
        current_row += 1


def write_instructions_to_sheet(
    worksheet: Worksheet, instructions: list[str], start_row: int, help_label_index: int, help_width: int
) -> int:
    current_row = start_row

    for instruction_index, instruction in enumerate(instructions):
        worksheet.cell(row=current_row, column=help_label_index, value=f"> {instruction_index + 1}.")
        instruction_lines = wrap_instructions(instruction, help_width)
        for line in instruction_lines:
            worksheet.cell(row=current_row, column=help_label_index + 1, value=line)
            current_row += 1

    return current_row


def read_column_header_titles(worksheet: Worksheet) -> list[str]:
    """Read the first row of the worksheet and return a list of these column titles."""
    index = 1
    titles: list[str] = []
    while True:
        value = worksheet.cell(1, index).value
        if not value:
            break
        titles.append(value)
        index += 1
    return titles
