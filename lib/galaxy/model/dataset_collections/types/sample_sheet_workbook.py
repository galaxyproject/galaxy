from dataclasses import dataclass
from typing import (
    cast,
    Dict,
    List,
    Optional,
    Protocol,
    Tuple,
    TYPE_CHECKING,
    Union,
)

from openpyxl import Workbook
from openpyxl.styles.protection import Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import (
    BaseModel,
    Field,
)
from typing_extensions import Literal

from galaxy.exceptions import RequestParameterInvalidException
from galaxy.model.dataset_collections.rule_target_columns import (
    column_titles_to_headers,
    HeaderColumn,
    InferredColumnMapping,
    ParsedColumn,
)
from galaxy.model.dataset_collections.rule_target_models import (
    ColumnTarget,
    COMMON_COLUMN_TARGETS,
    RuleBuilderMappingTargetKey,
    target_model_by_type,
)
from galaxy.model.dataset_collections.workbook_util import (
    add_extra_column_help_as_new_sheet,
    add_instructions_to_sheet,
    Base64StringT,
    ContentTypeMessage,
    CsvDialectInferenceMessage,
    ExtraColumnsHelpConfiguration,
    freeze_header_row,
    HasHelp,
    HelpConfiguration,
    index_to_excel_column,
    load_workbook_from_base64,
    make_headers_bold,
    parse_format_messages,
    ReadOnlyWorkbook,
    set_column_width,
)
from galaxy.schema.schema import SampleSheetColumnValueT
from galaxy.util import (
    string_as_bool,
    string_as_bool_or_none,
)
from .sample_sheet_util import SampleSheetColumnDefinitionModel

if TYPE_CHECKING:
    from galaxy.model import (
        DatasetCollection,
        DatasetCollectionElement,
    )


class DatasetCollectionElementLike(Protocol):
    id: int
    element_identifier: str


class DatasetCollectionLike(Protocol):
    id: int
    collection_type: str
    elements: List[DatasetCollectionElementLike]


# mypy doesn't recognize "str" and "Mapped[str]" as compatible type signatures,
# is there a better way to interface out these model objects?
AnyDatasetCollectionElement = Union["DatasetCollectionElement", DatasetCollectionElementLike]
AnyDatasetCollection = Union["DatasetCollection", DatasetCollectionLike]


DEFAULT_TITLE = "Sample Sheet for Galaxy"
URI_HELP = "The URL/URI for the target file."

PrefixRowValuesT = List[List[SampleSheetColumnValueT]]
InternalSampleSheetColumnValueT = Union[SampleSheetColumnValueT, "ModelObjectPrefixValue"]
InternalPrefixRowValuesT = List[List[InternalSampleSheetColumnValueT]]

CreateTitleField = Field(
    DEFAULT_TITLE,
    title="Title of the workbook to generate",
    description="A short title to give the workbook.",
)

ColumnDefinitionsField: List[SampleSheetColumnDefinitionModel] = Field(
    ...,
    title="Column Descriptions",
    description="A description of the columns expected in the workbook after the first columns described by 'prefix_columns_type'",
)

WorkbookContentField: Base64StringT = Field(
    ...,
    title="Workbook Content (Base 64 encoded)",
    description="The workbook content (the contents of the xlsx file) that have been base64 encoded.",
)
PrefixRowsField: Optional[PrefixRowValuesT] = Field(
    None,
    title="Prefix sample sheet values",
    description="An area to pre-populate URIs, etc...",
)

SampleSheetCollectionType = Literal[
    "sample_sheet", "sample_sheet:paired", "sample_sheet:paired_or_unpaired", "sample_sheet:record"
]

ParsedRow = Dict[str, SampleSheetColumnValueT]
ParsedRows = List[ParsedRow]

AnyLogMessage = Union[InferredColumnMapping, ContentTypeMessage, CsvDialectInferenceMessage]

SampleSheetParseLog = List[AnyLogMessage]


class ParsedWorkbook(BaseModel):
    rows: ParsedRows
    # extra columns contained in the supplied workbook that have relevant Galaxy metadata
    # maybe should be thought of as "suffix_columns" since they are after the prefix columns
    # and user-defined columns.
    extra_columns: List[ParsedColumn]
    parse_log: SampleSheetParseLog


class CreateWorkbookRequest(BaseModel):
    title: str = CreateTitleField
    collection_type: SampleSheetCollectionType
    prefix_columns_type: Literal["URI"] = "URI"
    column_definitions: List[SampleSheetColumnDefinitionModel] = ColumnDefinitionsField
    prefix_values: Optional[PrefixRowValuesT] = None


@dataclass
class CreateWorkbookRequestForCollection:
    title: str
    dataset_collection: AnyDatasetCollection
    column_definitions: List[SampleSheetColumnDefinitionModel] = ColumnDefinitionsField


class CreateWorkbook(BaseModel):
    title: str = CreateTitleField
    collection_type: SampleSheetCollectionType
    prefix_columns_type: Literal["URI", "ModelObjects"] = "URI"
    column_definitions: List[SampleSheetColumnDefinitionModel] = ColumnDefinitionsField
    prefix_values: Optional[InternalPrefixRowValuesT] = None


@dataclass
class CreateWorkbookForCollection:
    title: str
    dataset_collection: AnyDatasetCollection
    column_definitions: List[SampleSheetColumnDefinitionModel] = ColumnDefinitionsField


class ParseWorkbook(BaseModel):
    collection_type: SampleSheetCollectionType
    prefix_columns_type: Literal["URI", "ModelObjects"] = "URI"
    column_definitions: List[SampleSheetColumnDefinitionModel] = ColumnDefinitionsField
    content: str = WorkbookContentField


@dataclass
class ParseWorkbookForCollection:
    dataset_collection: AnyDatasetCollection
    column_definitions: List[SampleSheetColumnDefinitionModel] = ColumnDefinitionsField
    content: str = WorkbookContentField


AnyParseWorkbook = Union[ParseWorkbook, ParseWorkbookForCollection]


INSTRUCTIONS = [
    "Use this spreadsheet to describe your samples. For each sample (i.e. each file), ensure all the labeled columns are specified and correct.",
    "If you're using Google Sheets, data validation will be applied automatically - just make sure no cell values have a red mark indicating they are invalid.",
    "If you're using Microsft Excel, it is best to run data validation after you've completed filling out this sheet. This can be done by clicking on 'Data' > 'Data Validation' > 'Circle Invalid Data'.",
    "Once data entry is complete, drop this file back into Galaxy to finish creating a sample sheet collection for your inputs.",
]

# the first columns are very different based on what we're creating here, TODO write instructions
# for each collection type
INSTRUCTIONS_BY_COLLECTION_TYPE: Dict[SampleSheetCollectionType, List[str]] = cast(
    Dict[SampleSheetCollectionType, List[str]],
    {
        "sample_sheet": INSTRUCTIONS,
        "sample_sheet:paired": INSTRUCTIONS,
        "sample_sheet:paired_or_unpaired": INSTRUCTIONS,
        "sample_sheet:record": INSTRUCTIONS,
    },
)

EXTRA_COLUMN_INSTRUCTIONS = [
    "Extra metadata for the uploaded datasets can be specified by just adding columns with special headers to the sheet.",
    "These columns must be added *AFTER* the columns defined in the sample sheet.",
    "The list of column metadata type appears in this sheet and example column names are provided.",
]


def parse_workbook(payload: ParseWorkbook) -> ParsedWorkbook:
    workbook: ReadOnlyWorkbook = load_workbook_from_base64(payload.content)
    parse_log: SampleSheetParseLog = []
    parse_log.extend(parse_format_messages(workbook))
    extra_columns, inferred_columns_log = _read_extra_column_headers(workbook, payload)
    parse_log.extend(inferred_columns_log)
    rows = _load_row_data(workbook, payload, extra_columns)
    if not workbook.typed:
        _normalize_rows(rows, payload)
    return ParsedWorkbook(
        rows=rows,
        extra_columns=[c.parsed_column for c in extra_columns],
        parse_log=parse_log,
    )


def _normalize_rows(rows: ParsedRows, payload: AnyParseWorkbook) -> None:
    """Match column definition types to row values.

    The excel reader does not require this, it reads in typed values for both integers,
    floats, and booleans, but the csv reader does not do any of that and so we should
    validate all that here and normalize the expectations.

    This does not throw exceptions on validation errors it just fixes the types if they
    fix cleanly. Validation errors need to be uniform across both types of workbooks and
    this code does not apply to Excel.
    """
    column_definitions = payload.column_definitions
    for row in rows:
        for column_definition in column_definitions:
            column_name = column_definition.name
            if column_name not in row:
                continue
            value = row[column_name]
            if value is None and column_definition.optional:
                continue
            elif value is None:
                value = ""
            if value == "" and column_definition.optional:
                continue

            if column_definition.type == "int":
                try:
                    row[column_name] = int(value)
                except ValueError:
                    # TODO: capture this and log it.
                    pass
            elif column_definition.type == "float":
                try:
                    row[column_name] = float(value)
                except ValueError:
                    # TODO: capture this and log it.
                    pass
            elif column_definition.type == "boolean":
                if isinstance(value, bool):
                    continue
                if isinstance(value, str):
                    if column_definition.optional:
                        row[column_name] = string_as_bool_or_none(value)
                    else:
                        row[column_name] = string_as_bool(value)
                else:
                    # TODO: capture this and log it.
                    pass


def _read_extra_column_headers(
    workbook: ReadOnlyWorkbook, payload: AnyParseWorkbook
) -> Tuple[List[HeaderColumn], List[InferredColumnMapping]]:
    required_prefix_columns = prefix_columns(payload)
    required_column_names = [c.name for c in required_prefix_columns] + [c.name for c in payload.column_definitions]
    num_required_columns = len(required_column_names)
    column_titles = workbook.column_titles()
    extra_headers = column_titles[num_required_columns:]
    return column_titles_to_headers(extra_headers, column_offset=num_required_columns)


def parse_workbook_for_collection(payload: ParseWorkbookForCollection) -> ParsedWorkbook:
    workbook = load_workbook_from_base64(payload.content)
    parse_log: SampleSheetParseLog = []
    parse_log.extend(parse_format_messages(workbook))
    rows = _load_row_data(workbook, payload, [])
    return ParsedWorkbook(rows=rows, extra_columns=[], parse_log=parse_log)


# a base64 version of this so we can do short get URLs with real links in the API.
def generate_workbook_from_request(payload: CreateWorkbookRequest) -> Workbook:
    column_definitions = payload.column_definitions

    prefix_values = None
    if payload.prefix_values:
        prefix_values = payload.prefix_values

    create_object = CreateWorkbook(
        collection_type=payload.collection_type,
        title=payload.title,
        prefix_columns_type=payload.prefix_columns_type,
        prefix_values=prefix_values,
        column_definitions=column_definitions,
    )
    return generate_workbook(create_object)


def generate_workbook_from_request_for_collection(payload: CreateWorkbookRequestForCollection) -> Workbook:
    column_definitions = payload.column_definitions
    create_object = CreateWorkbookForCollection(
        title=payload.title,
        dataset_collection=payload.dataset_collection,
        column_definitions=column_definitions,
    )
    return generate_workbook_for_collection(create_object)


class ModelObjectPrefixValue(BaseModel):
    model_class: Literal["DatasetCollectionElement"]
    element_id: int
    element_identifier: str

    @staticmethod
    def from_dataset_collection_element(dce: AnyDatasetCollectionElement) -> "ModelObjectPrefixValue":
        return ModelObjectPrefixValue(
            model_class="DatasetCollectionElement",
            element_id=dce.id,
            element_identifier=dce.element_identifier,
        )


def generate_workbook(payload: CreateWorkbook) -> Workbook:
    prefix_column_types = payload.prefix_columns_type
    collection_type = payload.collection_type
    instructions = INSTRUCTIONS_BY_COLLECTION_TYPE[collection_type]

    # Create a workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = payload.title

    column_definitions = payload.column_definitions
    the_prefix_columns = prefix_columns(payload)
    num_initial_columns = len(the_prefix_columns)
    headers: List[HasHelp] = [c.has_help for c in the_prefix_columns] + [
        HasHelp(cd.name, cd.description or "") for cd in column_definitions
    ]
    worksheet.append([h.title for h in headers])
    make_headers_bold(worksheet, headers)

    for index, header in enumerate(headers):
        if "URI" in header.title:
            width = 80
        else:
            width = 20
        set_column_width(worksheet, index, width)

    _add_prefix_column_validations(payload, worksheet)
    freeze_header_row(worksheet)

    for index, column_definition in enumerate(column_definitions):
        validation: Optional[DataValidation] = None
        if column_definition.type == "int":
            validation = DataValidation(type="whole", allow_blank=True)
            # TODO: operator="between", formula1="1", formula2="1000"
        elif column_definition.type == "float":
            validation = DataValidation(type="decimal", allow_blank=True)
            # TODO: operator="between", formula1="1", formula2="1000"
        elif column_definition.type == "boolean":
            column_str = index_to_excel_column(index + num_initial_columns)
            validation = DataValidation(
                type="custom",
                formula1=f"OR({column_str}2=TRUE, {column_str}2=FALSE)",
                showDropDown=False,
                allow_blank=True,
            )

            dropdown_validation = DataValidation(
                type="list", formula1='"TRUE,FALSE"', showDropDown=False, allow_blank=True
            )
            _add_validation(index_to_excel_column(index + num_initial_columns), dropdown_validation, worksheet)
        elif column_definition.restrictions:
            list_as_formula = ",".join([str(r) for r in column_definition.restrictions])
            validation = DataValidation(
                type="list", formula1=f'"{list_as_formula}"', showDropDown=False, allow_blank=True
            )
            validation.prompt = "Please select from the list"

        if validation:
            validation.error = "Invalid input"
            validation.errorTitle = "Error"
            validation.promptTitle = column_definition.name
            _add_validation(index_to_excel_column(index + num_initial_columns), validation, worksheet)

    help_configuration = HelpConfiguration(
        instructions=instructions,
        columns=headers,
        text_width=50,
        column_width=50,
    )
    add_instructions_to_sheet(
        worksheet,
        help_configuration,
    )

    prefix_rows = payload.prefix_values or []
    prefix_rows_offset = 2  # header + 1-index-ed data structure
    for row_index, row in enumerate(prefix_rows):
        for column_index, col_value in enumerate(row):
            if isinstance(col_value, ModelObjectPrefixValue):
                col_value = col_value.element_identifier
            worksheet.cell(row=row_index + prefix_rows_offset, column=column_index + 1, value=col_value)

    if prefix_column_types == "ModelObjects":
        model_object_prefix_values = cast(List[ModelObjectPrefixValue], payload.prefix_values)
        _lock_sheet_for_existing_collection(worksheet, model_object_prefix_values, column_definitions)

        # Add another worksheet - is this what caused "corruption"?
        # additional_worksheet = workbook.create_sheet(title="Internal Galaxy Tracking (do not edit)")

    if prefix_column_types != "ModelObjects":
        extra_column_configuration = ExtraColumnsHelpConfiguration(
            EXTRA_COLUMN_INSTRUCTIONS, text_width=50, column_targets=COMMON_COLUMN_TARGETS
        )
        add_extra_column_help_as_new_sheet(workbook, extra_column_configuration)
    return workbook


def generate_workbook_for_collection(payload: CreateWorkbookForCollection) -> Workbook:
    input_collection_type = payload.dataset_collection.collection_type
    sample_sheet_collection_type = _list_to_sample_sheet_collection_type(input_collection_type)

    prefix_values: List[List[ModelObjectPrefixValue]] = []
    for element in payload.dataset_collection.elements:
        prefix_values.append([ModelObjectPrefixValue.from_dataset_collection_element(element)])

    create_workbook = CreateWorkbook(
        title=payload.title,
        collection_type=sample_sheet_collection_type,
        prefix_columns_type="ModelObjects",
        column_definitions=payload.column_definitions,
        prefix_values=prefix_values,
    )
    return generate_workbook(create_workbook)


@dataclass
class FetchPrefixColumn:
    type: RuleBuilderMappingTargetKey
    title: str  # user facing
    # e.g. for paired data will have two columns of URIs, record types maybe have any number
    # and after dataset hash may have multiples of those also
    type_index: int

    @property
    def name(self):
        if self.type_index == 0:
            return self.type
        else:
            return f"{self.type}_{self.type_index}"

    @property
    def help(self) -> str:
        column_target = _prefix_column_to_column_target(self)
        return column_target.help if column_target.help else ""

    @property
    def has_help(self):
        return HasHelp(title=self.title, help=self.title)


def prefix_columns(payload: Union[CreateWorkbook, AnyParseWorkbook]) -> List[FetchPrefixColumn]:
    if isinstance(payload, (CreateWorkbook, ParseWorkbook)):
        collection_type = payload.collection_type
        columns_type = payload.prefix_columns_type
    elif isinstance(payload, ParseWorkbookForCollection):
        list_collection_type = payload.dataset_collection.collection_type
        collection_type = _list_to_sample_sheet_collection_type(list_collection_type)
        columns_type = "ModelObjects"

    def uri_column(column_title: str, type_index: int = 0) -> FetchPrefixColumn:
        return FetchPrefixColumn(
            type="url",
            title=column_title,
            type_index=type_index,
        )

    def element_identifier_column() -> FetchPrefixColumn:
        return FetchPrefixColumn(
            type="list_identifiers",
            title="Element identifier",
            type_index=0,
        )

    if columns_type == "URI":
        if collection_type == "sample_sheet":
            columns = [uri_column("URI"), element_identifier_column()]
        elif collection_type == "sample_sheet:paired":
            columns = [uri_column("URI 1 (forward)"), uri_column("URI 2 (reverse)", 1), element_identifier_column()]
        elif collection_type == "sample_sheet:paired_or_unpaired":
            columns = [
                uri_column("URI 1 (forward if paired)"),
                uri_column("URI 2 (optional - reverse if paired)", 1),
                element_identifier_column(),
            ]
        else:
            raise NotImplementedError()
    elif columns_type == "ModelObjects":
        columns = [
            # override help?
            # "Element identifier of existing Galaxy collection, do not edit this value."
            FetchPrefixColumn(type="list_identifiers", title="Element Identifier", type_index=0)
        ]
    else:
        raise NotImplementedError("Unknown and unimplemented columns type encountered {columns_type}")
    return columns


def prefix_column_names(payload: Union[CreateWorkbook, AnyParseWorkbook]) -> List[str]:
    return [c.title for c in prefix_columns(payload)]


def prefix_column_counts(payload: CreateWorkbook) -> int:
    return len(prefix_columns(payload))


def _add_prefix_column_validations(payload: CreateWorkbook, worksheet: Worksheet):
    prefix_column_types = payload.prefix_columns_type
    if prefix_column_types == "URI":
        for i in range(prefix_column_counts(payload)):
            # Add data validation for "URI" column
            # We cannot assume http/https since drs, gxfiles, etc... are all fine
            uri_validation = DataValidation(type="custom", formula1='=ISNUMBER(FIND("://", E2))', allow_blank=True)
            uri_validation.error = "Invalid URI"
            uri_validation.errorTitle = "Error"
            uri_validation.showErrorMessage = True
            _add_validation(index_to_excel_column(i), uri_validation, worksheet)
    elif prefix_column_types == "ModelObjects":
        # these should be locked identifiers, no need to validate?
        # excel online prevents editing these but Google Sheets allows.
        pass
    else:
        raise NotImplementedError("Unknown and unimplemented columns type encountered {columns_type}")


def _lock_sheet_for_existing_collection(
    worksheet: Worksheet,
    prefix_values: List[ModelObjectPrefixValue],
    column_definitions: List[SampleSheetColumnDefinitionModel],
) -> None:
    worksheet.protection.sheet = True
    for column_prefix_index in range(len(column_definitions)):
        for row_prefix_index in range(len(prefix_values)):
            sheet_column_prefix_index = column_prefix_index + 1  # advance one column for the element identifier
            sheet_row_prefix_index = row_prefix_index + 1  # advance one column for column headers
            # advance one more for 1-based indices in format
            cell = worksheet.cell(sheet_row_prefix_index + 1, sheet_column_prefix_index + 1)
            cell.protection = Protection(locked=False)


def _add_validation(column: str, data_validation: DataValidation, worksheet: Worksheet):
    worksheet.add_data_validation(data_validation)
    data_validation.add(f"{column}2:{column}1048576")


def _load_row_data(
    workbook: ReadOnlyWorkbook, payload: AnyParseWorkbook, extra_columns: List[HeaderColumn]
) -> ParsedRows:
    rows: ParsedRows = []

    the_prefix_columns = prefix_columns(payload)
    column_names = [c.name for c in the_prefix_columns] + [c.name for c in payload.column_definitions]
    if extra_columns:
        column_names += [c.name for c in extra_columns]
    columns_to_read = len(column_names)
    for row_index, row in enumerate(workbook.iter_rows(columns_to_read)):
        if row_index == 0:  # skip column headers
            continue
        if not row[0]:
            break
        parsed_row: ParsedRow = {}
        for value, column_name in zip(row, column_names):
            parsed_row[column_name] = value
        rows.append(parsed_row)
    return rows


def _list_to_sample_sheet_collection_type(input_collection_type: str) -> SampleSheetCollectionType:
    """Convert simple list collection types to corresponding sample_sheet collection types.

    What would the sample_sheet collection type that allows decorating that kind of list. For instance,
    list:paired becomes sample_sheet:paired.
    """
    sample_sheet_collection_type: Optional[SampleSheetCollectionType] = None
    if input_collection_type == "list":
        sample_sheet_collection_type = "sample_sheet"
    elif input_collection_type == "list:paired":
        sample_sheet_collection_type = "sample_sheet:paired"
    elif input_collection_type == "list:paired_or_unpaired":
        sample_sheet_collection_type = "sample_sheet:paired_or_unpaired"
    elif input_collection_type == "list:record":
        raise NotImplementedError("Work in progress, this has not bee implemented yet")
    else:
        raise RequestParameterInvalidException(
            f"Invalid collection type for sample sheet workbook generation {input_collection_type}"
        )
    return sample_sheet_collection_type


def _prefix_column_to_column_target(column_header: FetchPrefixColumn) -> ColumnTarget:
    return target_model_by_type(column_header.type)
