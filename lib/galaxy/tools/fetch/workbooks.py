from dataclasses import dataclass
from typing import (
    Dict,
    List,
    Optional,
    Tuple,
    Union,
)

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import (
    BaseModel,
    Field,
)
from typing_extensions import Literal

from galaxy.exceptions import RequestParameterInvalidException
from galaxy.model.dataset_collections.auto_identifiers import (
    fill_in_identifiers,
    FillIdentifiers,
)
from galaxy.model.dataset_collections.auto_pairing import paired_element_list_identifier
from galaxy.model.dataset_collections.rule_target_columns import (
    column_titles_to_headers,
    HeaderColumn,
    implied_paired_or_unpaired_column_header,
    ParsedColumn,
)
from galaxy.model.dataset_collections.rule_target_models import (
    COMMON_COLUMN_TARGETS,
)
from galaxy.model.dataset_collections.workbook_util import (
    add_column_validation,
    add_extra_column_help_as_new_sheet,
    add_instructions_to_sheet,
    Base64StringT,
    ExtraColumnsHelpConfiguration,
    freeze_header_row,
    HasHelp,
    HelpConfiguration,
    load_workbook_from_base64,
    make_headers_bold,
    read_column_header_titles,
    set_column_width,
    uri_data_validation,
)

FetchWorkbookType = Literal["datasets", "collection", "collections"]
FetchWorkbookCollectionType = Literal["list", "list:paired", "list:list", "list:list:paired", "list:paired_or_unpaired"]

DEFAULT_FETCH_WORKBOOK_TYPE: FetchWorkbookType = "datasets"
DEFAULT_FETCH_WORKBOOK_COLLECTION_TYPE: FetchWorkbookCollectionType = "list"
DEFAULT_WORKBOOK_TITLE: str = "Galaxy Data Import"

INSTRUCTION_USE_THIS = "Use this spreadsheet to describe your datasets. For each dataset (i.e. each file), ensure all the labeled columns are specified and correct."
INSTRUCTION_EXTRA_COLUMNS = "Additional columns can be added if we you wish to specify additional metadata."  # mention second tab when implemented
INSTRUCTION_ONCE_COMPLETE_DATASETS = (
    "Once data entry is complete, drop this file back into Galaxy to finish creating datasets for your inputs."
)
INSTRUCTION_ONCE_COMPLETE_COLLECTION = (
    "Once data entry is complete, drop this file back into Galaxy to finish creating a collection for your inputs."
)
INSTRUCTION_ONCE_COMPLETE_COLLECTIONS = (
    "Once data entry is complete, drop this file back into Galaxy to finish creating collections for your inputs."
)
INSTRUCTIONS_BY_TYPE: Dict[FetchWorkbookType, List[str]] = {
    "datasets": [
        INSTRUCTION_USE_THIS,
        INSTRUCTION_EXTRA_COLUMNS,
        INSTRUCTION_ONCE_COMPLETE_DATASETS,
    ],
    "collection": [
        INSTRUCTION_USE_THIS,
        INSTRUCTION_EXTRA_COLUMNS,
        INSTRUCTION_ONCE_COMPLETE_COLLECTION,
    ],
    "collections": [
        INSTRUCTION_USE_THIS,
        INSTRUCTION_EXTRA_COLUMNS,
        INSTRUCTION_ONCE_COMPLETE_COLLECTIONS,
    ],
}

EXCEPTION_TOO_MANY_URI_COLUMNS = "Too many URI/URL columns in the supplied workbook."
EXCEPTION_NO_URIS_FOUND = "Failed to find any URI or URL like column in the supplied workbook."
EXTRA_COLUMN_INSTRUCTIONS = [
    "Extra metadata for the uploaded datasets can be specified by just adding columns with special headers to the first sheet of this workbook.",
    "The columns can appear in order on the first sheet of the workbook and the names are fairly flexible - extra white space and casing do not matter.",
    "The list of supported column metadata types appear in this help sheet to the left along with example column names Galaxy would support.",
]


@dataclass
class GenerateFetchWorkbookRequest:
    type: FetchWorkbookType = DEFAULT_FETCH_WORKBOOK_TYPE
    collection_type: FetchWorkbookCollectionType = DEFAULT_FETCH_WORKBOOK_COLLECTION_TYPE
    title: str = DEFAULT_WORKBOOK_TITLE


WorkbookContentField: Base64StringT = Field(
    ...,
    max_length=1024 * 1024 * 10,  # 10MB
    title="Workbook Content (Base 64 encoded)",
    description="The workbook content (the contents of the xlsx file) that have been base64 encoded.",
)


class ParseFetchWorkbook(BaseModel):
    content: Base64StringT = WorkbookContentField
    fill_identifiers: Optional[FillIdentifiers] = None


def header_column_to_parsed_column(header_column: HeaderColumn) -> "ParsedColumn":
    return header_column.parsed_column


def generate(request: GenerateFetchWorkbookRequest) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = request.title
    header_columns = _request_to_columns(request)

    header_titles = [h.title for h in header_columns]
    worksheet.append(header_titles)
    make_headers_bold(worksheet, header_titles)
    freeze_header_row(worksheet)
    for i, column in enumerate(header_columns):
        set_column_width(worksheet, i, column.width)

    uri_data_validator = uri_data_validation("A")
    add_column_validation("A", uri_data_validator, worksheet)

    columns_for_help = [HasHelp(c.title, c.help) for c in header_columns]
    help_configuration = HelpConfiguration(
        instructions=INSTRUCTIONS_BY_TYPE[request.type],
        columns=columns_for_help,
        text_width=50,
        column_width=50,
    )
    add_instructions_to_sheet(
        worksheet,
        help_configuration,
    )
    extra_column_configuration = ExtraColumnsHelpConfiguration(
        EXTRA_COLUMN_INSTRUCTIONS, text_width=50, column_targets=COMMON_COLUMN_TARGETS
    )
    add_extra_column_help_as_new_sheet(workbook, extra_column_configuration)
    return workbook


ParsedRow = Dict[str, Optional[str]]
ParsedRows = List[ParsedRow]


class ParseLogEntry(BaseModel):
    message: str


class SplitUpPairedDataLogEntry(ParseLogEntry):
    message: str
    old_forward_column: ParsedColumn
    old_reverse_column: ParsedColumn
    new_paired_status_column: int


class InferredCollectionTypeLogEntry(ParseLogEntry):
    message: str
    from_columns: List[ParsedColumn]


ParseLog = List[ParseLogEntry]


class BaseParsedFetchWorkbook(BaseModel):
    rows: ParsedRows
    columns: List[ParsedColumn]
    workbook_type: FetchWorkbookType
    parse_log: ParseLog


class ParsedFetchWorkbookForDatasets(BaseParsedFetchWorkbook):
    workbook_type: FetchWorkbookType = "datasets"


class ParsedFetchWorkbookForCollections(BaseParsedFetchWorkbook):
    workbook_type: FetchWorkbookType = "collection"
    collection_type: FetchWorkbookCollectionType


ParsedFetchWorkbook = Union[ParsedFetchWorkbookForDatasets, ParsedFetchWorkbookForCollections]


def parse(payload: ParseFetchWorkbook) -> ParsedFetchWorkbook:
    parse_log: ParseLog = []
    workbook = load_workbook_from_base64(payload.content)
    column_headers = _read_column_headers(workbook.active)
    _validate_parsed_column_headers(column_headers)
    raw_rows = _load_row_data(workbook, payload)
    # the rule builder does require splitting the paired data in this way but it might
    # be worth it to do it with an "initial rule" instead to demo how you'd do it
    # with actual rule builder rules? Not sure.
    rows, column_headers, split_data_log_entry = _split_paired_data_if_needed(raw_rows, column_headers)
    if split_data_log_entry:
        parse_log.append(split_data_log_entry)
    columns = [ParsedColumn(title=c.title, type=c.type, type_index=c.type_index) for c in column_headers]
    if _is_fetch_workbook_for_collections(column_headers):
        collection_type, log_entry = _infer_fetch_workbook_collection_type(column_headers)
        parse_log.append(log_entry)
        assert collection_type in ["list", "list:paired", "list:list", "list:list:paired", "list:paired_or_unpaired"]
        rows = _fill_in_identifier_column_if_needed(rows, column_headers, payload.fill_identifiers)
        return ParsedFetchWorkbookForCollections(
            collection_type=collection_type, rows=rows, columns=columns, parse_log=parse_log
        )
    else:
        return ParsedFetchWorkbookForDatasets(rows=rows, columns=columns, parse_log=parse_log)


def _validate_parsed_column_headers(column_headers: List[HeaderColumn]) -> None:
    uri_like_columns = _uri_like_columns(column_headers)
    if len(uri_like_columns) > 2:
        raise RequestParameterInvalidException(
            f"{EXCEPTION_TOO_MANY_URI_COLUMNS}. Relevant headers are {[c.title for c in uri_like_columns]}"
        )
    if len(uri_like_columns) == 0:
        raise RequestParameterInvalidException(EXCEPTION_NO_URIS_FOUND)


def _request_to_columns(request: GenerateFetchWorkbookRequest) -> List[HeaderColumn]:
    if request.type == "datasets":
        return [
            HeaderColumn("url", "URI", 0),
            HeaderColumn("name", "Name", 0),
        ]
    else:
        if request.collection_type == "list":
            header_columns = [
                HeaderColumn("url", "URI", 0),
                HeaderColumn("list_identifiers", "List Identifier", 0),
            ]
        elif request.collection_type == "list:paired":
            header_columns = [
                HeaderColumn("url", "URI 1 (Forward)", 0),
                HeaderColumn("url", "URI 2 (Reverse)", 1),
                HeaderColumn("list_identifiers", "List Identifier", 0),
            ]
        elif request.collection_type == "list:list":
            header_columns = [
                HeaderColumn("url", "URI", 0),
                HeaderColumn("list_identifiers", "Outer List Identifier", 0),
                HeaderColumn("list_identifiers", "Inner List Identifier", 1),
            ]
        elif request.collection_type == "list:list:paired":
            header_columns = [
                HeaderColumn("url", "URI 1 (Forward)", 0),
                HeaderColumn("url", "URI 2 (Reverse)", 1),
                HeaderColumn("list_identifiers", "Outer List Identifier", 0),
                HeaderColumn("list_identifiers", "Inner List Identifier", 1),
            ]
        elif request.collection_type == "list:paired_or_unpaired":
            header_columns = [
                HeaderColumn("url", "URI 1 (Forward)", 0),
                HeaderColumn("url", "URI 2 (Optional/Reverse)", 1),
                HeaderColumn("list_identifiers", "List Identifier", 0),
            ]
        else:
            raise NotImplementedError()
        if request.type == "collections":
            header_columns.append(
                HeaderColumn(
                    type="collection_name",
                    title="Collection Name",
                    type_index=0,
                )
            )
        return header_columns


def _load_row_data(workbook: Workbook, payload: ParseFetchWorkbook) -> ParsedRows:
    sheet = workbook.active  # Get the first sheet

    rows: ParsedRows = []

    column_headers = _read_column_headers(sheet)
    columns_to_read = len(column_headers)
    index_of_first_uri = _index_of_fist_uri_column(column_headers)

    for row_index, row in enumerate(sheet.iter_rows(max_col=columns_to_read, values_only=True)):
        if row_index == 0:  # skip column headers
            continue
        if not row[index_of_first_uri]:
            break
        parsed_row: ParsedRow = {}
        for value, column in zip(row, column_headers):
            parsed_row[column.name] = value
        rows.append(parsed_row)
    return rows


def _split_paired_data_if_needed(
    rows: ParsedRows, column_headers: List[HeaderColumn]
) -> Tuple[ParsedRows, List[HeaderColumn], Optional[SplitUpPairedDataLogEntry]]:
    split_rows: ParsedRows = []
    uri_like_columns = _uri_like_columns(column_headers)
    if len(_uri_like_columns(column_headers)) != 2:
        return rows, column_headers, None

    hash_columns_to_split: List[Tuple[HeaderColumn, HeaderColumn]] = []
    for column_type in ["hash_sha1", "hash_md5", "hash_sha256", "hash_sha512"]:
        hash_columns = [c for c in column_headers if c.type == column_type]
        if len(hash_columns) == 2:
            hash_columns_to_split.append((hash_columns[0], hash_columns[1]))

    url_column_0 = uri_like_columns[0]
    url_column_1 = uri_like_columns[1]

    # Split the data up
    split_rows = []
    for row in rows:
        row_to_copy = row.copy()

        url_0 = row_to_copy.pop(url_column_0.name)
        url_1 = row_to_copy.pop(url_column_1.name)

        # before we break up the pairs, do any implicit identifier filling in
        list_identifier_columns = [c for c in column_headers if c.type == "list_identifiers"]
        if len(list_identifier_columns) > 0:
            inner_list_identifier_column = list_identifier_columns[-1]
            list_identifier = row_to_copy[inner_list_identifier_column.name]
            if not list_identifier and url_0 and url_1:
                list_identifier = paired_element_list_identifier(url_0, url_1)
                row_to_copy[inner_list_identifier_column.name] = list_identifier

        row_0 = row_to_copy.copy()
        row_1 = row_to_copy.copy()
        row_0[url_column_0.name] = url_0
        row_1[url_column_0.name] = url_1
        row_0["paired_identifier"] = "1"
        row_1["paired_identifier"] = "2"

        for hash_column_0, hash_column_1 in hash_columns_to_split:
            # remove all the hashes from copied row before we...
            row_0.pop(hash_column_0.name, None)
            row_1.pop(hash_column_0.name, None)
            row_0.pop(hash_column_1.name, None)
            row_1.pop(hash_column_1.name, None)

            # ... merge both hash columns in the original into one column in the duplicates
            row_0[hash_column_0.name] = row.pop(hash_column_0.name)
            row_1[hash_column_0.name] = row.pop(hash_column_1.name)

        split_rows.append(row_0)
        split_rows.append(row_1)

    # Adjust the columns accordingly
    column_headers.remove(url_column_1)
    for _, hash_column_1 in hash_columns_to_split:
        column_headers.remove(hash_column_1)

    column_headers.append(
        HeaderColumn(
            type="paired_identifier",
            title="Paired Identifier",
            type_index=0,
        )
    )
    split_log_entry = SplitUpPairedDataLogEntry(
        message="Merged paired data URIs columns into single column and added paired identifier column.",
        old_forward_column=header_column_to_parsed_column(url_column_0),
        old_reverse_column=header_column_to_parsed_column(url_column_1),
        new_paired_status_column=len(column_headers) - 1,
    )
    return split_rows, column_headers, split_log_entry


def _fill_in_identifier_column_if_needed(
    rows: ParsedRows, columns: List[HeaderColumn], config: Optional[FillIdentifiers]
) -> ParsedRows:
    list_identifiers_columns = [c for c in columns if c.type == "list_identifiers"]
    uri_columns = _uri_like_columns(columns)
    if len(uri_columns) == 0 or len(list_identifiers_columns) == 0:
        # this doesn't look like a list creation, don't need identifiers to be filled in
        return rows
    if config is None or not config.fill_inner_list_identifiers:
        # no filling in identifiers requested, just return the rows as is
        return rows

    uri_column = uri_columns[0]
    inner_list_identifier_column = list_identifiers_columns[-1]
    uris_to_identifiers: List[Tuple[str, Optional[str]]] = []
    for row in rows:
        uri = row.get(uri_column.name)
        if not uri:
            continue
        list_identifier = row.get(inner_list_identifier_column.name)
        uris_to_identifiers.append((uri, list_identifier))
    new_identifiers = fill_in_identifiers(uris_to_identifiers, config)
    for row, new_identifier in zip(rows, new_identifiers):
        row[inner_list_identifier_column.name] = new_identifier
    return rows


def _read_column_headers(worksheet: Worksheet) -> List[HeaderColumn]:
    column_titles = read_column_header_titles(worksheet)
    return column_titles_to_headers(column_titles)


def _uri_like_columns(column_headers: List[HeaderColumn]) -> List[HeaderColumn]:
    return [c for c in column_headers if c.type == "url" or c.type == "url_deferred"]


def _index_of_fist_uri_column(column_headers: List[HeaderColumn]) -> int:
    for index, column_header in enumerate(column_headers):
        if column_header.type == "url" or column_header.type == "url_deferred":
            return index
    raise RequestParameterInvalidException(
        f"{EXCEPTION_NO_URIS_FOUND}. Relevant headers are {[c.title for c in column_headers]}"
    )


def _infer_fetch_workbook_collection_type(
    column_headers: List[HeaderColumn],
) -> Tuple[str, InferredCollectionTypeLogEntry]:
    paired_identifier_columns = [c for c in column_headers if c.type == "paired_identifier"]
    paired_or_unpaired_identifier_columns = [c for c in column_headers if c.type == "paired_or_unpaired_identifier"]
    any_paired = len(paired_identifier_columns) > 0
    uri_columns = _uri_like_columns(column_headers)
    num_uris = len(uri_columns)

    inference_on_columns: List[ParsedColumn] = []

    list_type: str = ""
    for column_header in column_headers:
        if column_header.type == "list_identifiers":
            inference_on_columns.append(header_column_to_parsed_column(column_header))
            if list_type:
                list_type = f"list:{list_type}"
            else:
                list_type = "list"
    if any_paired or num_uris == 2:
        if num_uris == 2:
            inference_on_columns.append(header_column_to_parsed_column(uri_columns[0]))
            inference_on_columns.append(header_column_to_parsed_column(uri_columns[1]))
            if implied_paired_or_unpaired_column_header(uri_columns[1]):
                collection_type = f"{list_type}:paired_or_unpaired"
            else:
                collection_type = f"{list_type}:paired"
        else:
            paired_identifier_column = paired_identifier_columns[0]
            inference_on_columns.append(header_column_to_parsed_column(paired_identifier_column))
            if implied_paired_or_unpaired_column_header(paired_identifier_column):
                collection_type = f"{list_type}:paired_or_unpaired"
            else:
                collection_type = f"{list_type}:paired"
    elif len(paired_or_unpaired_identifier_columns) > 0:
        inference_on_columns.append(header_column_to_parsed_column(paired_or_unpaired_identifier_columns[0]))
        collection_type = f"{list_type}:paired_or_unpaired"
    else:
        collection_type = list_type
    return collection_type, InferredCollectionTypeLogEntry(
        message="Inferred collection type from column headers.",
        from_columns=inference_on_columns,
    )


def _is_fetch_workbook_for_collections(column_headers: List[HeaderColumn]) -> bool:
    return _infer_fetch_workbook_collection_type(column_headers)[0] != ""
